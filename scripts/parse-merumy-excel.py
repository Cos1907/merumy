#!/usr/bin/env python3
import json
import sys
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    try:
        import openpyxl
        USE_OPENPYXL = True
    except ImportError:
        print("Error: Please install pandas or openpyxl: pip3 install pandas openpyxl")
        sys.exit(1)
    else:
        USE_OPENPYXL = True
        USE_PANDAS = False
else:
    USE_PANDAS = True
    USE_OPENPYXL = False

def parse_price(price_str):
    """Parse price string to float"""
    if not price_str or (isinstance(price_str, float) and pd.isna(price_str)):
        return None
    # Remove currency symbols and spaces, replace comma with dot
    price_str = str(price_str).strip().replace('TRY', '').replace(',', '.').replace('"', '').replace('₺', '').strip()
    try:
        return float(price_str)
    except:
        return None

def clean_text(text):
    """Clean text from extra spaces and newlines"""
    if not text or (isinstance(text, float) and pd.isna(text)):
        return ''
    return ' '.join(str(text).split())

def generate_slug(product_code, product_name):
    """Generate slug from product code and name"""
    slug = product_name.lower().replace(' ', '-').replace('--', '-')
    slug = ''.join(c for c in slug if c.isalnum() or c == '-')[:50]
    slug = f"{product_code}-{slug}"
    return slug

# Read Excel file
excel_path = Path(__file__).parent.parent / 'kbeauty' / 'Merumy E-ticaret.xlsx'

if not excel_path.exists():
    print(f"Error: Excel file not found at {excel_path}")
    sys.exit(1)

print(f"Reading Excel file: {excel_path}")

if USE_PANDAS:
    # Use pandas
    df = pd.read_excel(excel_path)
    print(f"Found {len(df)} rows")
    print(f"Columns: {df.columns.tolist()}")
    
    # Find column indices
    columns = df.columns.tolist()
    
    # Try to find the columns we need
    urun_kodu_col = None
    urun_adi_col = None
    barkod_col = None
    marka_col = None
    kategori_col = None  # Ozellik04Adi
    alt_kategori_col = None  # Ozellik05Adi
    fiyat_col = None
    gorsel_col = None
    
    for i, col in enumerate(columns):
        col_lower = str(col).lower()
        if 'urun' in col_lower and 'kod' in col_lower:
            urun_kodu_col = i
        elif 'urun' in col_lower and 'adi' in col_lower:
            urun_adi_col = i
        elif 'barkod' in col_lower:
            barkod_col = i
        elif 'marka' in col_lower:
            marka_col = i
        elif 'ozellik04adi' in col_lower or 'ozellik 04' in col_lower:
            kategori_col = i
        elif 'ozellik05adi' in col_lower or 'ozellik 05' in col_lower:
            alt_kategori_col = i
        elif 'fiyat' in col_lower or 'price' in col_lower:
            fiyat_col = i
        elif 'gorsel' in col_lower or 'image' in col_lower or 'resim' in col_lower:
            gorsel_col = i
    
    print(f"\nColumn mapping:")
    print(f"  UrunKodu: {urun_kodu_col} ({columns[urun_kodu_col] if urun_kodu_col is not None else 'NOT FOUND'})")
    print(f"  UrunAdi: {urun_adi_col} ({columns[urun_adi_col] if urun_adi_col is not None else 'NOT FOUND'})")
    print(f"  Barkod: {barkod_col} ({columns[barkod_col] if barkod_col is not None else 'NOT FOUND'})")
    print(f"  Marka: {marka_col} ({columns[marka_col] if marka_col is not None else 'NOT FOUND'})")
    print(f"  Kategori (Ozellik04Adi): {kategori_col} ({columns[kategori_col] if kategori_col is not None else 'NOT FOUND'})")
    print(f"  Alt Kategori (Ozellik05Adi): {alt_kategori_col} ({columns[alt_kategori_col] if alt_kategori_col is not None else 'NOT FOUND'})")
    print(f"  Fiyat: {fiyat_col} ({columns[fiyat_col] if fiyat_col is not None else 'NOT FOUND'})")
    print(f"  Görsel: {gorsel_col} ({columns[gorsel_col] if gorsel_col is not None else 'NOT FOUND'})")
    
    products = []
    categories = {}
    brands = {}
    
    for idx, row in df.iterrows():
        try:
            product_code = clean_text(row.iloc[urun_kodu_col]) if urun_kodu_col is not None else ''
            product_name = clean_text(row.iloc[urun_adi_col]) if urun_adi_col is not None else ''
            
            if not product_code or not product_name:
                continue
            
            barcode = clean_text(row.iloc[barkod_col]) if barkod_col is not None else ''
            brand_name = clean_text(row.iloc[marka_col]) if marka_col is not None else 'Merumy'
            category_name = clean_text(row.iloc[kategori_col]) if kategori_col is not None else 'Genel'
            subcategory_name = clean_text(row.iloc[alt_kategori_col]) if alt_kategori_col is not None else ''
            price = parse_price(row.iloc[fiyat_col]) if fiyat_col is not None else None
            image_link = clean_text(row.iloc[gorsel_col]) if gorsel_col is not None else ''
            
            if not price or price == 0:
                price = 0
            
            slug = generate_slug(product_code, product_name)
            
            product = {
                'id': product_code,
                'code': product_code,
                'slug': slug,
                'name': product_name,
                'brand': brand_name or 'Merumy',
                'category': category_name or 'Genel',
                'subcategory': subcategory_name or '',
                'price': price,
                'originalPrice': None,
                'image': image_link or '/images/product-placeholder.png',
                'barcode': barcode,
                'rating': round(4.0 + (abs(hash(product_code)) % 10) / 10, 1),
                'reviews': (abs(hash(product_code)) % 500) + 10,
                'sold': (abs(hash(product_code)) % 1000) + 50,
                'inStock': True,
                'description': f'{product_name} - {brand_name} markasından kaliteli {category_name} ürünü.'
            }
            
            products.append(product)
            
            # Track categories
            if category_name and category_name != 'Genel':
                if category_name not in categories:
                    categories[category_name] = {
                        'name': category_name,
                        'subcategories': set(),
                        'products': []
                    }
                categories[category_name]['products'].append(product)
                if subcategory_name:
                    categories[category_name]['subcategories'].add(subcategory_name)
            
            # Track brands
            if brand_name and brand_name != 'Merumy':
                if brand_name not in brands:
                    brands[brand_name] = []
                brands[brand_name].append(product)
                
        except Exception as e:
            print(f"Error processing row {idx}: {e}")
            continue

else:
    # Use openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # Read header row
    header_row = [cell.value for cell in ws[1]]
    print(f"Found {len(header_row)} columns")
    print(f"Columns: {header_row[:10]}...")
    
    # Find column indices
    urun_kodu_col = None
    urun_adi_col = None
    barkod_col = None
    marka_col = None
    kategori_col = None
    alt_kategori_col = None
    fiyat_col = None
    gorsel_col = None
    
    for i, col in enumerate(header_row):
        col_lower = str(col).lower() if col else ''
        if 'urun' in col_lower and 'kod' in col_lower:
            urun_kodu_col = i
        elif 'urun' in col_lower and 'adi' in col_lower:
            urun_adi_col = i
        elif 'barkod' in col_lower:
            barkod_col = i
        elif 'marka' in col_lower:
            marka_col = i
        elif 'ozellik04adi' in col_lower or 'ozellik 04' in col_lower:
            kategori_col = i
        elif 'ozellik05adi' in col_lower or 'ozellik 05' in col_lower:
            alt_kategori_col = i
        elif 'fiyat' in col_lower or 'price' in col_lower:
            fiyat_col = i
        elif 'gorsel' in col_lower or 'image' in col_lower or 'resim' in col_lower:
            gorsel_col = i
    
    print(f"\nColumn mapping:")
    print(f"  UrunKodu: {urun_kodu_col} ({header_row[urun_kodu_col] if urun_kodu_col is not None else 'NOT FOUND'})")
    print(f"  UrunAdi: {urun_adi_col} ({header_row[urun_adi_col] if urun_adi_col is not None else 'NOT FOUND'})")
    print(f"  Barkod: {barkod_col} ({header_row[barkod_col] if barkod_col is not None else 'NOT FOUND'})")
    print(f"  Marka: {marka_col} ({header_row[marka_col] if marka_col is not None else 'NOT FOUND'})")
    print(f"  Kategori (Ozellik04Adi): {kategori_col} ({header_row[kategori_col] if kategori_col is not None else 'NOT FOUND'})")
    print(f"  Alt Kategori (Ozellik05Adi): {alt_kategori_col} ({header_row[alt_kategori_col] if alt_kategori_col is not None else 'NOT FOUND'})")
    print(f"  Fiyat: {fiyat_col} ({header_row[fiyat_col] if fiyat_col is not None else 'NOT FOUND'})")
    print(f"  Görsel: {gorsel_col} ({header_row[gorsel_col] if gorsel_col is not None else 'NOT FOUND'})")
    
    products = []
    categories = {}
    brands = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if len(row) < max(filter(None, [urun_kodu_col, urun_adi_col, barkod_col, marka_col, kategori_col, alt_kategori_col, fiyat_col, gorsel_col])):
                continue
            
            product_code = clean_text(row[urun_kodu_col]) if urun_kodu_col is not None and urun_kodu_col < len(row) else ''
            product_name = clean_text(row[urun_adi_col]) if urun_adi_col is not None and urun_adi_col < len(row) else ''
            
            if not product_code or not product_name:
                continue
            
            barcode = clean_text(row[barkod_col]) if barkod_col is not None and barkod_col < len(row) else ''
            brand_name = clean_text(row[marka_col]) if marka_col is not None and marka_col < len(row) else 'Merumy'
            category_name = clean_text(row[kategori_col]) if kategori_col is not None and kategori_col < len(row) else 'Genel'
            subcategory_name = clean_text(row[alt_kategori_col]) if alt_kategori_col is not None and alt_kategori_col < len(row) else ''
            price = parse_price(row[fiyat_col]) if fiyat_col is not None and fiyat_col < len(row) else None
            image_link = clean_text(row[gorsel_col]) if gorsel_col is not None and gorsel_col < len(row) else ''
            
            if not price or price == 0:
                price = 0
            
            slug = generate_slug(product_code, product_name)
            
            product = {
                'id': product_code,
                'code': product_code,
                'slug': slug,
                'name': product_name,
                'brand': brand_name or 'Merumy',
                'category': category_name or 'Genel',
                'subcategory': subcategory_name or '',
                'price': price,
                'originalPrice': None,
                'image': image_link or '/images/product-placeholder.png',
                'barcode': barcode,
                'rating': round(4.0 + (abs(hash(product_code)) % 10) / 10, 1),
                'reviews': (abs(hash(product_code)) % 500) + 10,
                'sold': (abs(hash(product_code)) % 1000) + 50,
                'inStock': True,
                'description': f'{product_name} - {brand_name} markasından kaliteli {category_name} ürünü.'
            }
            
            products.append(product)
            
            # Track categories
            if category_name and category_name != 'Genel':
                if category_name not in categories:
                    categories[category_name] = {
                        'name': category_name,
                        'subcategories': set(),
                        'products': []
                    }
                categories[category_name]['products'].append(product)
                if subcategory_name:
                    categories[category_name]['subcategories'].add(subcategory_name)
            
            # Track brands
            if brand_name and brand_name != 'Merumy':
                if brand_name not in brands:
                    brands[brand_name] = []
                brands[brand_name].append(product)
                
        except Exception as e:
            print(f"Error processing row {row_idx}: {e}")
            continue

# Convert sets to lists for JSON
for cat in categories.values():
    cat['subcategories'] = list(cat['subcategories'])

# Save products data
output_dir = Path(__file__).parent.parent / 'data'
output_dir.mkdir(exist_ok=True)

# Save all products
with open(output_dir / 'products.json', 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

# Save categories
categories_data = {k: {
    'name': v['name'],
    'subcategories': v['subcategories'],
    'productCount': len(v['products'])
} for k, v in categories.items()}

with open(output_dir / 'categories.json', 'w', encoding='utf-8') as f:
    json.dump(categories_data, f, ensure_ascii=False, indent=2)

# Save brands
brands_data = {k: len(v) for k, v in brands.items()}
with open(output_dir / 'brands.json', 'w', encoding='utf-8') as f:
    json.dump(brands_data, f, ensure_ascii=False, indent=2)

print(f"\n✅ Parsed {len(products)} products")
print(f"✅ Found {len(categories)} categories")
print(f"✅ Found {len(brands)} brands")
print(f"\nTop categories:")
for cat_name, cat_data in sorted(categories_data.items(), key=lambda x: x[1]['productCount'], reverse=True)[:10]:
    print(f"  - {cat_name}: {cat_data['productCount']} products")



